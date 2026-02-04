"""
Excel Reader with Multilingual Support
Handles English, Hindi, and Marathi text in Excel files
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging
from typing import Dict, List, Any, Optional
import re

logger = logging.getLogger(__name__)


class ExcelReader:
    """
    Reads Excel files with support for multilingual content (English, Hindi, Marathi)
    """
    
    # Expected column headers (flexible matching)
    EXPECTED_HEADERS = [
        'Sr.',
        'Budget Item No.',
        'District',
        'Head of Accounts',
        'Name of the work',
        'Name Of The Work (In Marathi)',
        'Administrative Approval Cost (Lakh)',
        'Administrative Approval Date',
        'Technical Sanction Cost (Lakh)',
        'Contract Agreement Cost (Lakh)',
        'Conract % above / below',
        'Expenditure Upto March (Lakhs)',
        'Expenditure April to January (Lakhs)',
        'Total Expenditure (Lakhs)',
        '% of Expenditure',
        'Physical Progress',
        'Date of Work_Order',
        'Original Time Limit in Days',
        'Work Category',
        'Road Category',
        'Chainage From',
        'Chainage To',
        'Taluka',
        'Constituency',
        'PW Division',
        'PW Circle',
        'PW Region',
        'User Department'
    ]
    
    def __init__(self, encoding: str = 'utf-8'):
        self.encoding = encoding
        self.df = None
        self.raw_data = None
        self.column_mapping = {}
        
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Read Excel file with multilingual support
        
        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (default: first sheet)
            
        Returns:
            DataFrame with cleaned and validated data
        """
        logger.info(f"Reading Excel file: {file_path}")
        
        try:
            # Try reading with openpyxl (better Unicode support)
            wb = load_workbook(file_path, read_only=False, data_only=True)
            
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Read all data
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            # Convert to DataFrame
            if len(data) > 0:
                self.df = pd.DataFrame(data[1:], columns=data[0])
            else:
                raise ValueError("Excel file is empty")
            
            logger.info(f"Successfully read {len(self.df)} rows and {len(self.df.columns)} columns")
            
            # Clean and validate
            self.df = self._clean_data(self.df)
            self._validate_columns()
            self.df = self._standardize_data_types()
            
            return self.df
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            
            # Fallback to pandas
            try:
                logger.info("Attempting fallback with pandas read_excel")
                self.df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
                self.df = self._clean_data(self.df)
                self._validate_columns()
                self.df = self._standardize_data_types()
                return self.df
            except Exception as e2:
                logger.error(f"Fallback also failed: {e2}")
                raise Exception(f"Could not read Excel file: {e}")
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean the raw data"""
        logger.info("Cleaning data...")
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Clean column names
        df.columns = [self._clean_column_name(col) for col in df.columns]
        
        # Remove duplicate headers (sometimes Excel has repeated header rows)
        df = self._remove_duplicate_headers(df)
        
        # Trim whitespace from string columns
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        logger.info(f"Data cleaned: {len(df)} rows remaining")
        return df
    
    def _clean_column_name(self, col_name: str) -> str:
        """Clean column name while preserving multilingual characters"""
        if pd.isna(col_name):
            return "Unnamed"
        
        col_name = str(col_name).strip()
        # Remove excessive whitespace
        col_name = re.sub(r'\s+', ' ', col_name)
        return col_name
    
    def _remove_duplicate_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove rows that are duplicate headers"""
        # Check if any row matches the header
        header_row = df.columns.tolist()
        
        rows_to_drop = []
        for idx, row in df.iterrows():
            row_values = row.tolist()
            # Check if row matches header
            if self._is_header_row(row_values, header_row):
                rows_to_drop.append(idx)
        
        if rows_to_drop:
            logger.info(f"Removing {len(rows_to_drop)} duplicate header rows")
            df = df.drop(rows_to_drop)
        
        return df.reset_index(drop=True)
    
    def _is_header_row(self, row_values: List, header: List) -> bool:
        """Check if a row is a duplicate header"""
        if len(row_values) != len(header):
            return False
        
        matches = 0
        for rv, hv in zip(row_values, header):
            if pd.isna(rv) and pd.isna(hv):
                matches += 1
            elif str(rv).strip().lower() == str(hv).strip().lower():
                matches += 1
        
        # If more than 80% match, it's likely a header row
        return matches / len(header) > 0.8
    
    def _validate_columns(self):
        """Validate that expected columns are present"""
        logger.info("Validating columns...")
        
        missing_critical = []
        
        # Critical columns that must be present
        critical_columns = [
            'Budget Item No.',
            'Name of the work',
            'Administrative Approval Cost (Lakh)',
            'Total Expenditure (Lakhs)'
        ]
        
        current_columns = self.df.columns.tolist()
        
        for critical_col in critical_columns:
            # Fuzzy match
            found = False
            for current_col in current_columns:
                if self._columns_match(critical_col, current_col):
                    found = True
                    self.column_mapping[critical_col] = current_col
                    break
            
            if not found:
                missing_critical.append(critical_col)
        
        if missing_critical:
            logger.warning(f"Missing critical columns: {missing_critical}")
            logger.info(f"Available columns: {current_columns}")
        else:
            logger.info("All critical columns found")
        
        # Map all expected columns
        for expected_col in self.EXPECTED_HEADERS:
            if expected_col not in self.column_mapping:
                for current_col in current_columns:
                    if self._columns_match(expected_col, current_col):
                        self.column_mapping[expected_col] = current_col
                        break
    
    def _columns_match(self, expected: str, actual: str, threshold: float = 0.8) -> bool:
        """
        Check if two column names match (fuzzy matching)
        """
        expected_clean = expected.lower().replace(' ', '').replace('_', '').replace('.', '')
        actual_clean = actual.lower().replace(' ', '').replace('_', '').replace('.', '')
        
        # Exact match
        if expected_clean == actual_clean:
            return True
        
        # Check if one is substring of other
        if expected_clean in actual_clean or actual_clean in expected_clean:
            return True
        
        # Simple similarity check
        matches = sum(1 for a, b in zip(expected_clean, actual_clean) if a == b)
        max_len = max(len(expected_clean), len(actual_clean))
        
        if max_len == 0:
            return False
        
        similarity = matches / max_len
        return similarity >= threshold
    
    def _standardize_data_types(self) -> pd.DataFrame:
        """Standardize data types for analysis"""
        logger.info("Standardizing data types...")
        
        df = self.df.copy()
        
        # Numeric columns
        numeric_columns = [
            'Administrative Approval Cost (Lakh)',
            'Technical Sanction Cost (Lakh)',
            'Contract Agreement Cost (Lakh)',
            'Expenditure Upto March (Lakhs)',
            'Expenditure April to January (Lakhs)',
            'Total Expenditure (Lakhs)',
            '% of Expenditure',
            'Physical Progress',
            'Original Time Limit in Days',
            'Chainage From',
            'Chainage To',
            'Conract % above / below'
        ]
        
        for col in numeric_columns:
            actual_col = self.column_mapping.get(col, col)
            if actual_col in df.columns:
                df[actual_col] = pd.to_numeric(df[actual_col], errors='coerce')
        
        # Date columns
        date_columns = [
            'Administrative Approval Date',
            'Date of Work_Order'
        ]
        
        for col in date_columns:
            actual_col = self.column_mapping.get(col, col)
            if actual_col in df.columns:
                df[actual_col] = pd.to_datetime(df[actual_col], errors='coerce')
        
        # Ensure standard column names for easier processing
        rename_dict = {v: k for k, v in self.column_mapping.items() if v in df.columns}
        df = df.rename(columns=rename_dict)
        
        logger.info("Data types standardized")
        return df
    
    def get_data_summary(self) -> Dict[str, Any]:
        """Get summary of the loaded data"""
        if self.df is None:
            return {"error": "No data loaded"}
        
        return {
            'total_rows': len(self.df),
            'total_columns': len(self.df.columns),
            'columns': self.df.columns.tolist(),
            'null_counts': self.df.isnull().sum().to_dict(),
            'data_types': self.df.dtypes.astype(str).to_dict(),
            'sample_data': self.df.head(3).to_dict('records')
        }
    
    def validate_data_quality(self) -> Dict[str, Any]:
        """Validate data quality"""
        if self.df is None:
            return {"error": "No data loaded"}
        
        issues = []
        
        # Check for missing critical values
        critical_cols = [
            'Budget Item No.',
            'Name of the work',
            'Administrative Approval Cost (Lakh)'
        ]
        
        for col in critical_cols:
            if col in self.df.columns:
                null_count = self.df[col].isnull().sum()
                if null_count > 0:
                    issues.append({
                        'type': 'missing_critical_data',
                        'column': col,
                        'count': int(null_count),
                        'percentage': round(null_count / len(self.df) * 100, 2)
                    })
        
        # Check for suspicious values
        if 'Total Expenditure (Lakhs)' in self.df.columns:
            negative_exp = (self.df['Total Expenditure (Lakhs)'] < 0).sum()
            if negative_exp > 0:
                issues.append({
                    'type': 'negative_expenditure',
                    'count': int(negative_exp)
                })
        
        return {
            'total_issues': len(issues),
            'issues': issues,
            'quality_score': max(0, 100 - len(issues) * 5)
        }
